from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, auth

import os, uuid
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook

router = APIRouter()

UPLOAD_DIR = "app/uploads/rapports"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@router.get("/rapports/pdf")
def export_pdf(db: Session = Depends(get_db), current_user = Depends(auth.require_role(["admin"]))):
    budgets = db.query(models.Budget).all()
    depenses = db.query(models.Depense).all()

    file_path = os.path.join(UPLOAD_DIR, f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Rapport Institutionnel - Budgets et Dépenses", styles["Title"]))
    elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')} par {current_user.username}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Tableau Budgets
    data_budgets = [["ID", "Nom", "Montant (FCFA)", "Prévision (FCFA)"]]
    for b in budgets:
        data_budgets.append([b.id, b.nom, b.montant, b.prevision or "-"])

    table_budgets = Table(data_budgets, colWidths=[50, 150, 100, 100])
    table_budgets.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 12),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Budgets", styles["Heading2"]))
    elements.append(table_budgets)
    elements.append(Spacer(1, 30))

    # Tableau Dépenses (❌ sans colonne Facture)
    data_depenses = [["ID", "Description", "Montant (FCFA)", "Budget", "Statut"]]
    for d in depenses:
        data_depenses.append([d.id, d.description, d.montant, d.budget_id, d.statut])

    table_depenses = Table(data_depenses, colWidths=[40, 150, 100, 60, 100])
    table_depenses.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#006699")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 12),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Dépenses", styles["Heading2"]))
    elements.append(table_depenses)

    doc.build(elements)

    return FileResponse(file_path, media_type="application/pdf", filename="rapport.pdf")

@router.get("/rapports/excel")
def export_excel(db: Session = Depends(get_db), current_user = Depends(auth.require_role(["admin"]))):
    budgets = db.query(models.Budget).all()
    depenses = db.query(models.Depense).all()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Budgets"

    ws1.append(["ID", "Nom", "Montant", "Prévision"])
    for b in budgets:
        ws1.append([b.id, b.nom, b.montant, b.prevision])

    ws2 = wb.create_sheet("Dépenses")
    ws2.append(["ID", "Description", "Montant", "Budget", "Statut"])
    for d in depenses:
        ws2.append([d.id, d.description, d.montant, d.budget_id, d.statut])

    # ✅ Feuille Résumé
    ws3 = wb.create_sheet("Résumé")
    total_budgets = sum(b.montant for b in budgets)
    total_depenses_validees = sum(d.montant for d in depenses if d.statut in ["validé_controleur", "approuvé_admin"])
    pourcentage_global = (total_depenses_validees / total_budgets * 100) if total_budgets > 0 else 0

    ws3.append(["Total Budgets", total_budgets])
    ws3.append(["Total Dépenses Validées", total_depenses_validees])
    ws3.append(["Pourcentage Global Consommé", f"{round(pourcentage_global, 2)} %"])

    file_path = os.path.join(UPLOAD_DIR, f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(file_path)

    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="rapport.xlsx")
